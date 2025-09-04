#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel转换程序 v2.0
功能：
1. 将所有单元格高度改为140磅
2. 将所有单元格设置为垂直和水平居中
3. 在H列右侧插入新一列为I列
4. 将I列列宽改为42字符
5. 将H列URL转换为图片并嵌入I列单元格
6. 隐藏H列
7. 检测I列内容摘要字数，小于100字标红
8. 检测G列演员，含"-"横杠标红
9. 根据T列初审人匹配V列身份证号后四位
10. 人员库管理功能
"""

import sys
import os
import subprocess
import importlib
from typing import Optional, Dict, List
import time
import json
import re

# ============================================================================
#                           微剧URL转换工具v1.6
# ============================================================================
# 功能：将H列URL转换为图片并嵌入H列单元格
# 支持格式：.xlsx, .xls
# 新增功能：内容检测、人员库管理
# Bug反馈：李浩林
# 版本：1.6
# ============================================================================

# 人员库文件路径
STAFF_DB_FILE = "staff_database.json"
# 配置文件路径（位于脚本同目录）
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(BASE_DIR, "config.ini")

def print_banner():
    """打印程序横幅"""
    print("\n" + "="*80)
    print(" "*20 + "微剧URL转换工具v1.6" + " "*20)
    print(" "*15 + "功能：将H列URL转换为图片并嵌入H列单元格" + " "*15)
    print(" "*25 + "支持格式：.xlsx, .xls" + " "*25)
    print(" "*20 + "新增：内容检测、人员库管理" + " "*20)
    print(" "*30 + "Bug反馈：李浩林" + " "*30)
    print(" "*35 + "版本：1.6" + " "*35)
    print("="*80 + "\n")

def print_step_header(step_name, step_number, total_steps):
    """打印步骤头部"""
    print(f"\n{'='*60}")
    print(f"📋 步骤 {step_number}/{total_steps}: {step_name}")
    print(f"{'='*60}")

def print_step_complete(step_name):
    """打印步骤完成信息"""
    print(f"✅ {step_name} 完成！")
    print(f"{'─'*60}\n")

def print_progress_bar(current, total, description="", bar_length=50):
    """显示美化的进度条"""
    percentage = int((current / total) * 100)
    filled_length = int(bar_length * current // total)
    bar = '█' * filled_length + '░' * (bar_length - filled_length)
    
    # 添加动画效果
    animation_chars = ['⠋', '⠙', '⠹', '⠸', '⠼', '⠴', '⠦', '⠧', '⠇', '⠏']
    animation = animation_chars[int(time.time() * 10) % len(animation_chars)]
    
    print(f'\r{animation} {description} [{bar}] {percentage:3d}% ({current}/{total})', end='', flush=True)
    if current == total:
        print()  # 换行

def print_status(message, status_type="info"):
    """打印状态信息"""
    icons = {
        "info": "ℹ️",
        "success": "✅",
        "warning": "⚠️",
        "error": "❌",
        "loading": "🔄"
    }
    icon = icons.get(status_type, "ℹ️")
    print(f"{icon} {message}")

def check_and_install_package(package_name, install_name=None):
    """检查并安装缺少的包"""
    if install_name is None:
        install_name = package_name
    
    try:
        importlib.import_module(package_name)
        return True
    except ImportError:
        print_status(f"检测到缺少 {package_name} 库", "warning")
        print_status(f"正在安装 {install_name}...", "loading")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", install_name])
            print_status(f"{install_name} 安装成功", "success")
            return True
        except subprocess.CalledProcessError:
            print_status(f"{install_name} 安装失败", "error")
            return False

def install_required_packages():
    """安装所需的包"""
    print_step_header("检查运行库", 2, 6)
    
    packages = [
        ("PIL", "Pillow"),
        ("requests", "requests"),
        ("openpyxl", "openpyxl")
    ]
    
    missing_packages = []
    
    for i, (package_name, install_name) in enumerate(packages, 1):
        print_progress_bar(i, len(packages), f"检查 {package_name}")
        if not check_and_install_package(package_name, install_name):
            missing_packages.append(install_name)
    
    if missing_packages:
        print(f"\n❌ 以下包安装失败: {', '.join(missing_packages)}")
        print("请手动安装后重新运行程序")
        try:
            input("按回车键退出...")
        except (EOFError, RuntimeError):
            pass
        sys.exit(1)
    
    print_step_complete("运行库检查")

# 打印程序横幅
print_banner()

# 模式选择
def select_mode():
    """选择处理模式并保存到配置文件"""
    import configparser
    config = configparser.ConfigParser()

    # 尝试从配置文件读取模式
    try:
        config.read(CONFIG_FILE)
        saved_mode = config.get("DEFAULT", "mode", fallback=None)
        if saved_mode in ("1", "2"):
            print(f"\n当前模式为: 模式{saved_mode}")
    except Exception as e:
        print(f"读取配置文件失败: {e}")
    
    print("\n请选择处理模式：")
    print("1. 模式1：按原逻辑50条拆分")
    print("2. 模式2：不进行拆分")
    
    while True:
        mode = input("请输入模式编号(1/2): ").strip()
        if mode in ("1", "2"):
            # 保存模式到配置文件
            try:
                config["DEFAULT"] = {"mode": mode}
                with open(CONFIG_FILE, "w") as f:
                    config.write(f)
            except Exception as e:
                print(f"保存模式到配置文件失败: {e}")
            return int(mode)
        print("输入无效，请重新输入！")


def load_mode_config() -> int:
    """从配置文件读取模式设置"""
    import configparser

    config = configparser.ConfigParser()
    try:
        config.read(CONFIG_FILE)
        mode_str = config.get("DEFAULT", "mode", fallback="1")
        if mode_str in ("1", "2"):
            return int(mode_str)
    except Exception:
        pass
    return 1

# 安装所需的包
install_required_packages()

# 现在导入所需的库
print_step_header("导入运行库", 3, 6)
print_status("正在导入必要的库...", "loading")
import requests
from PIL import Image
import io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
print_step_complete("运行库导入")

def copy_cell_style(source_cell, target_cell):
    """复制单元格样式"""
    if source_cell.font:
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            color=source_cell.font.color
        )
    if source_cell.fill:
        target_cell.fill = PatternFill(
            fill_type=source_cell.fill.fill_type,
            start_color=source_cell.fill.start_color,
            end_color=source_cell.fill.end_color
        )
    if source_cell.border:
        target_cell.border = Border(
            left=source_cell.border.left,
            right=source_cell.border.right,
            top=source_cell.border.top,
            bottom=source_cell.border.bottom
        )
    if source_cell.alignment:
        target_cell.alignment = Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical,
            wrap_text=source_cell.alignment.wrap_text,
            shrink_to_fit=source_cell.alignment.shrink_to_fit
        )

def copy_sheet_formatting(source_sheet, target_sheet):
    """复制工作表格式设置"""
    # 复制列宽
    for col in range(1, source_sheet.max_column + 1):
        col_letter = get_column_letter(col)
        if col_letter in source_sheet.column_dimensions:
            target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width
    # 复制行高
    for row in range(1, source_sheet.max_row + 1):
        if row in source_sheet.row_dimensions:
            target_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height

def print_header():
    """打印程序头部界面（保留兼容性）"""
    print_banner()

def is_valid_url(url):
    if not url or not isinstance(url, str):
        return False
    url_pattern = re.compile(
        r'^https?://'
        r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+[A-Z]{2,6}\.?|'
        r'localhost|'
        r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})'
        r'(?::\d+)?'
        r'(?:/?|[/?]\S+)$', re.IGNORECASE)
    return bool(url_pattern.match(url.strip()))

def download_image(url):
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        content_type = response.headers.get('content-type', '').lower()
        if not content_type.startswith('image/'):
            return None
        return response.content
    except Exception as e:
        print(f"下载图片失败 {url}: {e}")
        return None

def print_progress(current, total, description=""):
    """显示进度条"""
    percentage = int((current / total) * 100)
    bar_length = 40
    filled_length = int(bar_length * current // total)
    bar = '█' * filled_length + '░' * (bar_length - filled_length)
    print(f'\r{description} [{bar}] {percentage}% ({current}/{total})', end='', flush=True)
    if current == total:
        print()  # 换行

def set_cell_alignment(sheet):
    """设置所有单元格为垂直和水平居中"""
    print_status("设置单元格对齐方式...", "loading")
    total_cells = sheet.max_row * sheet.max_column
    processed_cells = 0
    
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            processed_cells += 1
            if processed_cells % 100 == 0:  # 每100个单元格更新一次进度
                print_progress_bar(processed_cells, total_cells, "设置单元格对齐")
    
    print_progress_bar(total_cells, total_cells, "设置单元格对齐")
    print_step_complete("单元格对齐设置")

def embed_image_in_cell(sheet, row, col, image_path):
    """将图片嵌入到指定单元格中"""
    try:
        # 创建图片对象
        img = OpenpyxlImage(image_path)
        
        # 设置图片尺寸为单元格大小
        img.width = int(42 * 8)  # 42字符宽度
        img.height = int(140 * 1.33)  # 140磅高度
        
        # 将图片插入到指定单元格
        sheet.add_image(img, f'{get_column_letter(col)}{row}')
        
        # 设置单元格对齐方式
        cell = sheet.cell(row=row, column=col)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        return True
    except Exception as e:
        print(f"嵌入图片失败: {e}")
        return False

def convert_urls_to_images(workbook):
    total_sheets = len(workbook.worksheets)
    temp_files = []  # 存储临时文件路径
    
    print_step_header("处理Excel文件", 4, 6)
    
    # 加载人员库
    staff_db = load_staff_database()
    
    for sheet_idx, sheet in enumerate(workbook.worksheets, 1):
        print(f"\n📊 处理工作表 {sheet_idx}/{total_sheets}: {sheet.title}")
        
        h_column = 8
        g_column = 7  # 演员列
        i_column = 9  # 内容摘要列
        t_column = 20  # 初审人列
        v_column = 22  # 身份证号列
        
        # 设置所有单元格为垂直和水平居中
        set_cell_alignment(sheet)
        
        # 设置所有行高
        print_status("设置行高...", "loading")
        for row in range(1, sheet.max_row + 1):
            sheet.row_dimensions[row].height = 140
        print_step_complete("行高设置")
        
        # 内容检测和人员匹配
        print_status("进行内容检测和人员匹配...", "loading")
        content_check_count = 0
        actor_check_count = 0
        staff_match_count = 0
        
        for row in range(2, sheet.max_row + 1):
            # 检测I列内容摘要字数
            i_cell = sheet.cell(row=row, column=i_column)
            if i_cell.value:
                content_length = check_content_length(i_cell.value)
                if content_length < 100:
                    # 设置红色背景
                    from openpyxl.styles import PatternFill
                    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
                    i_cell.fill = red_fill
                    content_check_count += 1
            
            # 检测G列演员姓名
            g_cell = sheet.cell(row=row, column=g_column)
            if g_cell.value and check_actor_name(g_cell.value):
                # 设置红色背景
                from openpyxl.styles import PatternFill
                red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
                g_cell.fill = red_fill
                actor_check_count += 1
            
            # 根据T列初审人匹配V列身份证号
            t_cell = sheet.cell(row=row, column=t_column)
            v_cell = sheet.cell(row=row, column=v_column)
            if t_cell.value and staff_db:
                matched_id = match_staff_id(t_cell.value, staff_db)
                if matched_id:
                    v_cell.value = matched_id
                    staff_match_count += 1
        
        print_status(f"内容检测完成: I列标红{content_check_count}个, G列标红{actor_check_count}个, 人员匹配{staff_match_count}个", "success")
        
        # 统计需要处理的URL数量，并收集所有图片
        print_status("扫描URL...", "loading")
        url_count = 0
        url_rows = []
        url_list = []
        for row in range(2, sheet.max_row + 1):  # H1保留
            h_cell = sheet.cell(row=row, column=h_column)
            if h_cell.value and is_valid_url(str(h_cell.value)):
                url_count += 1
                url_rows.append(row)
                url_list.append(str(h_cell.value).strip())
        
        if url_count == 0:
            print_status("未找到有效的URL", "warning")
            continue
        
        print_status(f"找到 {url_count} 个URL需要处理", "info")
        processed_count = 0
        success_count = 0
        image_paths: list[Optional[str]] = [None for _ in range(url_count)]
        
        # 先批量下载图片
        print_status("开始下载图片...", "loading")
        for idx, url in enumerate(url_list):
            processed_count += 1
            print_progress_bar(processed_count, url_count, f"🖼️  下载图片 {processed_count}/{url_count}")
            image_data = download_image(url)
            if image_data:
                try:
                    pil_image = Image.open(io.BytesIO(image_data))
                    # 目标尺寸
                    target_width = int(42 * 8)   # 336
                    target_height = int(140 * 1.33)  # 186
                    # 缩放图片
                    pil_image = pil_image.convert('RGB')  # 保证JPEG兼容
                    # 兼容Pillow 10+和旧版本的缩放参数
                    try:
                        from PIL import Image as PILImage
                        resample_method = PILImage.Resampling.LANCZOS
                    except (ImportError, AttributeError):
                        resample_method = getattr(Image, 'LANCZOS', getattr(Image, 'BICUBIC', 3))
                    pil_image = pil_image.resize((target_width, target_height), resample_method)
                    # 保存为JPEG，压缩质量80
                    temp_path = f"temp_image_{url_rows[idx]}_{int(time.time())}.jpg"
                    pil_image.save(temp_path, format='JPEG', quality=80)
                    temp_files.append(temp_path)
                    image_paths[idx] = temp_path
                except Exception as e:
                    print_status(f"处理图片失败: {e}", "error")
                    if 'temp_path' in locals() and temp_path is not None and os.path.exists(temp_path):
                        os.remove(temp_path)
            else:
                print_status(f"无法下载图片: {url}", "error")
        print()  # 保证进度条换行
        
        # 清空H列（除H1）
        print_status("清空H列内容...", "loading")
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=h_column).value = None
        print_step_complete("H列清空")
        
        # 插入图片到H列
        print_status("插入图片到H列...", "loading")
        for idx, row in enumerate(url_rows):
            temp_path = image_paths[idx]
            if temp_path is not None and os.path.exists(temp_path):
                try:
                    img = OpenpyxlImage(temp_path)
                    img.width = int(42 * 8)  # 42字符宽度
                    img.height = int(140 * 1.33)  # 140磅高度
                    sheet.add_image(img, f'{get_column_letter(h_column)}{row}')
                    cell = sheet.cell(row=row, column=h_column)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    success_count += 1
                except Exception as e:
                    print_status(f"嵌入图片失败: {e}", "error")
        
        print_status(f"工作表 {sheet.title} 处理完成: 成功 {success_count}/{url_count}", "success")
    
    print_step_complete("Excel文件处理")
    return temp_files  # 返回临时文件列表

def convert_urls_to_images_and_split(workbook, excel_file):
    from typing import Optional
    from openpyxl import Workbook
    base_name = os.path.splitext(excel_file)[0]
    sheet = workbook.active
    h_column = 8
    g_column = 7  # 演员列
    i_column = 9  # 内容摘要列
    t_column = 20  # 初审人列
    v_column = 22  # 身份证号列
    max_row = sheet.max_row
    max_col = sheet.max_column
    temp_files = []
    
    # 加载人员库
    staff_db = load_staff_database()
    
    # 读取首行（表头）
    header = [sheet.cell(row=1, column=col).value for col in range(1, max_col + 1)]
    # 收集所有数据行和URL
    data_rows = []
    url_rows = []
    url_list = []
    for row in range(2, max_row + 1):
        row_data = [sheet.cell(row=row, column=col).value for col in range(1, max_col + 1)]
        data_rows.append(row_data)
        h_cell = sheet.cell(row=row, column=h_column)
        if h_cell.value and is_valid_url(str(h_cell.value)):
            url_rows.append(row - 2)  # 对应data_rows的索引
            url_list.append(str(h_cell.value).strip())
    total_data = len(data_rows)
    # 批量下载图片
    print_status(f"共{total_data}条数据，开始批量下载图片...", "info")
    image_paths: list[Optional[str]] = [None for _ in range(total_data)]
    for idx, url_idx in enumerate(url_rows):
        url = url_list[idx]
        print_progress_bar(idx + 1, len(url_list), f"🖼️  下载图片 {idx + 1}/{len(url_list)}")
        image_data = download_image(url)
        if image_data:
            try:
                pil_image = Image.open(io.BytesIO(image_data))
                target_width = int(42 * 8)
                target_height = int(140 * 1.33)
                pil_image = pil_image.convert('RGB')
                # 兼容Pillow 10+和旧版本的缩放参数
                try:
                    from PIL import Image as PILImage
                    resample_method = PILImage.Resampling.LANCZOS
                except (ImportError, AttributeError):
                    resample_method = getattr(Image, 'LANCZOS', getattr(Image, 'BICUBIC', 3))
                pil_image = pil_image.resize((target_width, target_height), resample_method)
                temp_path = f"temp_image_{url_idx+2}_{int(time.time())}.jpg"
                pil_image.save(temp_path, format='JPEG', quality=80)
                temp_files.append(temp_path)
                image_paths[url_idx] = temp_path
            except Exception as e:
                print_status(f"处理图片失败: {e}", "error")
                if 'temp_path' in locals() and temp_path is not None and os.path.exists(temp_path):
                    os.remove(temp_path)
        else:
            print_status(f"无法下载图片: {url}", "error")
    print()
    
    # 拆分数据
    print_status("开始拆分文件...", "loading")
    part_size = 50
    part_count = (total_data + part_size - 1) // part_size
    output_files = []
    
    for part in range(part_count):
        print_status(f"处理第 {part+1}/{part_count} 个文件...", "loading")
        
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = sheet.title if sheet and hasattr(sheet, 'title') else 'Sheet1'
        
        # 复制工作表格式设置
        copy_sheet_formatting(sheet, ws)
        
        # 写入表头并复制样式
        for col in range(1, max_col + 1):
            ws.cell(row=1, column=col, value=header[col-1])
            # 复制原表首行对应列的样式
            source_cell = sheet.cell(row=1, column=col)
            target_cell = ws.cell(row=1, column=col)
            copy_cell_style(source_cell, target_cell)
        
        # 写入数据
        for i in range(part_size):
            data_idx = part * part_size + i
            if data_idx >= total_data:
                break
            for col in range(1, max_col + 1):
                ws.cell(row=i+2, column=col, value=data_rows[data_idx][col-1])
        
        # 设置样式和行高
        set_cell_alignment(ws)
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 140
        ws.column_dimensions[get_column_letter(h_column)].width = 42
        
        # 内容检测和人员匹配
        print_status("进行内容检测和人员匹配...", "loading")
        content_check_count = 0
        actor_check_count = 0
        staff_match_count = 0
        
        for row in range(2, ws.max_row + 1):
            # 检测I列内容摘要字数
            i_cell = ws.cell(row=row, column=i_column)
            if i_cell.value:
                content_length = check_content_length(i_cell.value)
                if content_length < 100:
                    # 设置红色背景
                    from openpyxl.styles import PatternFill
                    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
                    i_cell.fill = red_fill
                    content_check_count += 1
            
            # 检测G列演员姓名
            g_cell = ws.cell(row=row, column=g_column)
            if g_cell.value and check_actor_name(g_cell.value):
                # 设置红色背景
                from openpyxl.styles import PatternFill
                red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
                g_cell.fill = red_fill
                actor_check_count += 1
            
            # 根据T列初审人匹配V列身份证号
            t_cell = ws.cell(row=row, column=t_column)
            v_cell = ws.cell(row=row, column=v_column)
            if t_cell.value and staff_db:
                matched_id = match_staff_id(t_cell.value, staff_db)
                if matched_id:
                    v_cell.value = matched_id
                    staff_match_count += 1
        
        print_status(f"内容检测完成: I列标红{content_check_count}个, G列标红{actor_check_count}个, 人员匹配{staff_match_count}个", "success")
        
        # 清空H列（除H1）
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=h_column).value = None
        
        # 插入图片
        for i in range(part_size):
            data_idx = part * part_size + i
            if data_idx >= total_data:
                break
            temp_path = image_paths[data_idx]
            if temp_path is not None and os.path.exists(temp_path):
                try:
                    img = OpenpyxlImage(temp_path)
                    img.width = int(42 * 8)
                    img.height = int(140 * 1.33)
                    ws.add_image(img, f'{get_column_letter(h_column)}{i+2}')
                    cell = ws.cell(row=i+2, column=h_column)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                except Exception as e:
                    print_status(f"嵌入图片失败: {e}", "error")
        
        # 保存文件
        if part_count == 1:
            output_file = f"{base_name}_converted.xlsx"
        else:
            output_file = f"{base_name}_part{part+1}.xlsx"
        
        wb.save(output_file)
        output_files.append(output_file)
        print_status(f"已保存: {output_file}", "success")
    return temp_files, output_files

# copy_h1_to_i1_and_delete_h_column 不再需要，可保留但不调用

def load_staff_database():
    """加载人员库"""
    if os.path.exists(STAFF_DB_FILE):
        try:
            with open(STAFF_DB_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print_status(f"加载人员库失败: {e}", "error")
            return create_default_staff_database()
    else:
        # 如果文件不存在，创建默认人员库
        return create_default_staff_database()

def create_default_staff_database():
    """创建默认人员库"""
    default_staff = {
        "梁应伟": "001X",
        "邹林伶": "5829",
        "赵志强": "7299",
        "杨华": "4241",
        "廖政": "1610",
        "万亭": "174X",
        "任雪梅": "5802",
        "冉小娟": "1363",
        "张静": "8525"
    }
    
    # 保存默认人员库
    if save_staff_database(default_staff):
        print_status("已创建默认审核人员库", "success")
        return default_staff
    else:
        print_status("创建默认人员库失败", "error")
        return {}

def save_staff_database(staff_db):
    """保存人员库"""
    try:
        with open(STAFF_DB_FILE, 'w', encoding='utf-8') as f:
            json.dump(staff_db, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print_status(f"保存人员库失败: {e}", "error")
        return False

def extract_chinese_name(name):
    """提取中文姓名"""
    if not name:
        return ""
    # 匹配中文字符，排除横杠等符号
    chinese_chars = re.findall(r'[\u4e00-\u9fff]+', str(name))
    return ''.join(chinese_chars)

def staff_management_menu():
    """人员库管理菜单"""
    print("\n" + "="*60)
    print(" "*15 + "👥 人员库管理系统" + " "*15)
    print("="*60)
    
    staff_db = load_staff_database()
    
    while True:
        print("\n📋 当前人员库:")
        if staff_db:
            for name, id_last4 in staff_db.items():
                print(f"   👤 {name} - 身份证后四位: {id_last4}")
        else:
            print("   📭 人员库为空")
        
        print("\n🔧 操作选项:")
        print("   1. 录入新人员")
        print("   2. 删除人员")
        print("   3. 返回主程序")
        
        try:
            choice = input("\n请选择操作 (1-3): ").strip()
            
            if choice == "1":
                add_new_staff(staff_db)
            elif choice == "2":
                delete_staff(staff_db)
            elif choice == "3":
                print_status("返回主程序", "info")
                break
            else:
                print_status("无效选择，请重新输入", "warning")
        except (EOFError, KeyboardInterrupt):
            print_status("返回主程序", "info")
            break

def add_new_staff(staff_db):
    """添加新人员"""
    print("\n📝 录入新人员信息")
    print("-" * 40)
    
    try:
        name = input("请输入人员姓名: ").strip()
        if not name:
            print_status("姓名不能为空", "error")
            return
        
        id_last4 = input("请输入身份证后四位: ").strip()
        if not id_last4 or len(id_last4) != 4 or not id_last4.isdigit():
            print_status("身份证后四位必须是4位数字", "error")
            return
        
        staff_db[name] = id_last4
        if save_staff_database(staff_db):
            print_status(f"人员 {name} 录入成功", "success")
        else:
            print_status("保存失败", "error")
    except (EOFError, KeyboardInterrupt):
        print_status("取消录入", "info")

def delete_staff(staff_db):
    """删除人员"""
    if not staff_db:
        print_status("人员库为空，无需删除", "info")
        return
    
    print("\n🗑️  删除人员")
    print("-" * 40)
    
    try:
        name = input("请输入要删除的人员姓名: ").strip()
        if name in staff_db:
            del staff_db[name]
            if save_staff_database(staff_db):
                print_status(f"人员 {name} 删除成功", "success")
            else:
                print_status("保存失败", "error")
        else:
            print_status(f"未找到人员 {name}", "warning")
    except (EOFError, KeyboardInterrupt):
        print_status("取消删除", "info")

def match_staff_id(name, staff_db):
    """根据姓名匹配身份证后四位"""
    if not name or not staff_db:
        return None
    
    chinese_name = extract_chinese_name(name)
    if not chinese_name:
        return None
    
    return staff_db.get(chinese_name)

def check_content_length(text):
    """检查文本长度"""
    if not text:
        return 0
    return len(str(text).strip())

def check_actor_name(name):
    """检查演员姓名是否包含横杠"""
    if not name:
        return False
    return "-" in str(name)

def main():
    print_step_header("程序启动", 1, 6)
    
    # 检查是否有命令行参数
    if len(sys.argv) < 2:
        # 没有参数，显示主菜单
        show_main_menu()
        return
    
    excel_file = sys.argv[1]
    if not os.path.exists(excel_file):
        print_status(f"文件不存在: {excel_file}", "error")
        try:
            input("按回车键退出...")
        except (EOFError, RuntimeError):
            pass
        return
    
    if not excel_file.lower().endswith(('.xlsx', '.xls')):
        print_status("请选择Excel文件 (.xlsx 或 .xls)", "error")
        try:
            input("按回车键退出...")
        except (EOFError, RuntimeError):
            pass
        return
    
    print_step_complete("程序启动")
    
    temp_files = []  # 存储临时文件路径
    
    try:
        print_step_header("文件处理", 5, 6)
        print_status(f"正在处理文件: {excel_file}", "info")
        print_status("加载工作簿...", "loading")
        
        # 尝试加载工作簿
        try:
            workbook = load_workbook(excel_file, data_only=True)
        except Exception as e:
            print_status(f"加载工作簿失败: {e}", "error")
            print_status("尝试使用兼容模式加载...", "loading")
            try:
                workbook = load_workbook(excel_file, keep_vba=False)
            except Exception as e2:
                print_status(f"兼容模式也失败: {e2}", "error")
                try:
                    input("按回车键退出...")
                except (EOFError, RuntimeError):
                    pass
                return
        
        # 判断是否需要拆分
        sheet = workbook.active
        if sheet is None:
            print_status("无法获取活动工作表", "error")
            return
            
        data_row_count = sheet.max_row - 1
        if mode == 1 and data_row_count > 50:
            print_status(f"数据行数超过50，按模式1拆分为多个文件...", "warning")
            temp_files, output_files = convert_urls_to_images_and_split(workbook, excel_file)
        else:
            print_status("按模式2处理，不进行拆分...", "info")
            temp_files = convert_urls_to_images(workbook)
            base_name = os.path.splitext(excel_file)[0]
            output_files = [f"{base_name}_converted.xlsx"]
            # 保存单一文件
            workbook.save(output_files[0])
        
        print_step_header("完成处理", 6, 6)
        print_status("图片已批量插入H列，请在Excel中手动设置图片为'嵌入单元格'（右键图片→设置属性→移动并调整大小）。", "warning")
        print_status("保存文件...", "loading")
        
        for output_file in output_files:
            print("\n" + "="*80)
            print(" "*25 + "🎉 转换完成！ 🎉" + " "*25)
            print(" "*20 + f"输出文件: {output_file}" + " "*20)
            print(" "*30 + "Bug反馈：李浩林" + " "*30)
            print("="*80)
            print("\033[1;33;41m【重要提示】点击图片Ctrl+A全选后，右键选择嵌入单元格（移动并调整大小）\033[0m")
        
        print_step_complete("文件处理")
        
    except Exception as e:
        print_status(f"处理文件时出错: {e}", "error")
    finally:
        # 清理临时文件
        print_status("清理临时文件...", "loading")
        for temp_file in temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
            except Exception as e:
                print_status(f"删除临时文件失败 {temp_file}: {e}", "warning")
    
    try:
        input("按回车键退出...")
    except (EOFError, RuntimeError):
        pass

# 全局模式变量
mode = load_mode_config()  # 从配置文件读取模式，默认模式1

def show_main_menu():
    """显示主菜单"""
    global mode
    while True:
        print("\n" + "="*60)
        print(" "*20 + "🎯 主菜单" + " "*20)
        print("="*60)
        print(f"\n当前模式: 模式{mode}")
        print("\n📋 功能选项:")
        print("   1. 审核人员身份证录入")
        print("   2. 拆分模式选择")
        print("   3. 处理Excel文件（拖拽文件到此程序）")
        print("   4. 退出程序")
        
        try:
            choice = input("\n请选择功能 (1-4): ").strip()
            
            if choice == "1":
                staff_management_menu()
            elif choice == "2":
                mode = select_mode()
            elif choice == "3":
                print_status("请将Excel文件拖拽到此程序上", "info")
                try:
                    input("按回车键返回主菜单...")
                except (EOFError, RuntimeError):
                    pass
            elif choice == "4":
                print_status("程序退出", "info")
                break
            else:
                print_status("无效选择，请重新输入", "warning")
        except (EOFError, KeyboardInterrupt):
            print_status("程序退出", "info")
            break

if __name__ == "__main__":
    main()